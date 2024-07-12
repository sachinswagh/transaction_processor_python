import os
import openpyxl
from tabulate import tabulate

from customer import Customer
from opening import Opening
from credit import Credit
from debit import Debit

cwd = os.getcwd()
path = f"{cwd}\day_3_assignment_input.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

cust_transactions = {}
customers = {
  1: Customer(1),
  2: Customer(2)
}

BOTTOM_ROWS_COUNT = 5

# ['Txnid', 'TxDate', 'TxType', 'Acct_id', 'Amount']
header = [i.value for i in list(sheet_obj.rows)[0]]
display_header = ['Acct_ID', 'Txn_Date', 'Txn_Amount', 'Closing Bal']

def get_cell_value(i, h):
  col = header.index(h)+1
  val = sheet_obj.cell(row = i, column = col).value
  return(val)

def create_txn(id, txn_date, txn_type, account_id, amount, customer):
  if txn_type == "O":
    return(Opening(id, txn_date, txn_type, account_id, amount, customer))
  if txn_type == "D":
    return(Debit(id, txn_date, txn_type, account_id, amount, customer))
  if txn_type == "C":
    return(Credit(id, txn_date, txn_type, account_id, amount, customer))

def process():
  for i in range(2,sheet_obj.max_row-BOTTOM_ROWS_COUNT):
    id = get_cell_value(i, 'Txnid')
    txn_date = get_cell_value(i, 'TxDate')
    txn_type = get_cell_value(i, 'TxType')
    account_id = get_cell_value(i, 'Acct_id')
    amount = get_cell_value(i, 'Amount')

    customer = customers[account_id]
    txn = create_txn(id, txn_date, txn_type, account_id, amount, customer)
    if account_id not in cust_transactions.keys():
      cust_transactions[account_id] = []
    cust_transactions[account_id].append(txn)

def sort_txn(txn):
  order = ['O', 'D', 'C']
  return(int(txn.txn_date.timestamp()), order.index(txn.txn_type) )

def sort_transactions(cust_transactions):
  for c_id in cust_transactions:
    transactions = cust_transactions[c_id]
    transactions.sort(key=sort_txn)
    cust_transactions[c_id] = transactions

def process_transactions(cust_transactions):
  for c_id in cust_transactions:
    customer = customers[c_id]
    for txn in cust_transactions[c_id]:
      txn.operate(customer)

def get_txn_row(txn):
  return(list([txn.account_id, txn.txn_date.date(), txn.signed_amount(), txn.closing_balance]))

def print_passbook(cust_transactions, cust_id):
  rows = [get_txn_row(txn) for txn in cust_transactions[cust_id]]
  print(tabulate(rows, display_header, tablefmt="pretty"))


process()
sort_transactions(cust_transactions)
process_transactions(cust_transactions)

for c_id in customers:
  print_passbook(cust_transactions, c_id)
  print("\n")
